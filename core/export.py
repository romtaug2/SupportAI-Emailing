"""
core/export.py
---------------

Export depuis SQLite vers CSV / XLSX / JSON pour chaque vertical. Les
colonnes et noms de fichiers sont identiques à ceux des notebooks d'origine,
pour que la sortie reste parfaitement compatible avec tes process en aval.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STRIPE_FILL = PatternFill("solid", fgColor="EAF2F8")
EMAIL_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_csv(
    rows: Iterable[dict],
    columns: list[str],
    path: Path,
) -> int:
    """
    Écrit un CSV UTF-8 BOM (compatible Excel français) avec exactement les
    colonnes demandées, dans l'ordre. Les clés absentes sont remplies à "".
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({c: row.get(c, "") or "" for c in columns})
            n += 1
    return n


def export_jsonl(rows: Iterable[dict], path: Path) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")
            n += 1
    return n


def export_xlsx(
    rows: list[dict],
    columns: list[str],
    path: Path,
    sheet_name: str = "Base",
    table_name: str = "BaseProspection",
    email_column: str | None = None,
    summary: list[tuple[str, object]] | None = None,
) -> int:
    """
    Écrit un fichier XLSX stylé (en-tête bleu, bordures, zébrage, figé ligne 1,
    auto-filtre, table nommée). Le style est volontairement identique à tes
    notebooks.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)

    for row in rows:
        ws.append([row.get(c, "") or "" for c in columns])

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row_cells[0].row % 2 == 0:
            for cell in row_cells:
                cell.fill = STRIPE_FILL

    # Largeurs auto (plafonnées)
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    if ws.max_row >= 2 and ws.max_column >= 1:
        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # Mise en valeur de la colonne email si précisée
    if email_column and email_column in columns and ws.max_row >= 2:
        idx = columns.index(email_column) + 1
        letter = get_column_letter(idx)
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            if cell.value and str(cell.value).strip():
                cell.fill = EMAIL_FILL

    # Feuille "Résumé" si demandée
    if summary:
        s = wb.create_sheet("Résumé", 0)
        s["A1"] = "Résumé export"
        s["A1"].font = Font(size=16, bold=True, color="1F4E78")
        for i, (label, value) in enumerate(summary, start=3):
            s[f"A{i}"] = label
            s[f"B{i}"] = value
            s[f"A{i}"].font = Font(bold=True)
            s[f"A{i}"].fill = PatternFill("solid", fgColor="F3F4F6")
            s[f"A{i}"].border = BORDER
            s[f"B{i}"].border = BORDER
        s.column_dimensions["A"].width = 32
        s.column_dimensions["B"].width = 48

    wb.save(path)
    return ws.max_row - 1
